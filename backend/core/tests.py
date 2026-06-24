import importlib
from io import BytesIO
from types import SimpleNamespace

from django.apps import apps as django_apps
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase, APIRequestFactory

from .data_io import IMPORT_KEY_HEADER, SHEET_CONFIG, export_data, import_data
from .models import (
    ClassSubjectTeacher, CombinedClassGroup, ScheduleLock, SchedulerSettings,
    SchoolClass, Subject, Teacher, TeacherBlockedTime, TeacherQualification,
)


class ImportDataTests(APITestCase):
    def setUp(self):
        self.factory = APIRequestFactory()

    def build_workbook(self, sheet_rows):
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, rows in sheet_rows.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(SHEET_CONFIG[sheet_name]['headers'])
            for row in rows:
                ws.append(row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_import_rejects_row_when_foreign_key_lookup_fails(self):
        school_class = SchoolClass.objects.create(name='一1班', grade=1)
        subject = Subject.objects.create(name='语文')
        schedule_lock_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is ScheduleLock
        )

        workbook_bytes = self.build_workbook({
            schedule_lock_sheet: [
                [school_class.name, subject.name, '不存在的教师', 1, 2],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(ScheduleLock.objects.count(), 0)
        self.assertFalse(response.data['committed'])
        self.assertEqual(response.data['results'][schedule_lock_sheet]['created'], 0)
        self.assertEqual(response.data['results'][schedule_lock_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)

    def test_import_teacher_combined_class_day(self):
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )
        combined_group = CombinedClassGroup.objects.create(name='校本一组')

        workbook_bytes = self.build_workbook({
            teacher_sheet: [
                ['李老师', None, combined_group.name, 3, 'FALSE', None, None],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 200)
        teacher = Teacher.objects.get(name='李老师')
        self.assertEqual(teacher.combined_class_group, combined_group)
        self.assertEqual(teacher.combined_class_day, 3)

    def test_import_rolls_back_all_rows_when_any_error_occurs(self):
        school_class = SchoolClass.objects.create(name='二1班', grade=2)
        subject = Subject.objects.create(name='数学')
        teacher = Teacher.objects.create(name='张老师')
        schedule_lock_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is ScheduleLock
        )

        workbook_bytes = self.build_workbook({
            schedule_lock_sheet: [
                [school_class.name, subject.name, teacher.name, 1, 2],
                [school_class.name, subject.name, '不存在的教师', 1, 3],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(ScheduleLock.objects.count(), 0)
        self.assertEqual(response.data['results'][schedule_lock_sheet]['created'], 0)
        self.assertEqual(response.data['results'][schedule_lock_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)

    def test_import_rejects_assignment_for_class_meeting_subject(self):
        school_class = SchoolClass.objects.create(name='三1班', grade=3)
        teacher = Teacher.objects.create(name='班主任老师')
        class_meeting_name = SchedulerSettings.get_settings().class_meeting_name
        subject = Subject.objects.create(name=class_meeting_name)
        assignment_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is ClassSubjectTeacher
        )

        workbook_bytes = self.build_workbook({
            assignment_sheet: [
                [school_class.name, subject.name, teacher.name, 'TRUE'],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(ClassSubjectTeacher.objects.count(), 0)
        self.assertEqual(response.data['results'][assignment_sheet]['created'], 0)
        self.assertEqual(response.data['results'][assignment_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('班会和校本课程不在授课分配中设置。', response.data['errors'][0])

    def test_import_rejects_assignment_for_inapplicable_grade_subject(self):
        school_class = SchoolClass.objects.create(name='四1班', grade=4)
        teacher = Teacher.objects.create(name='信息老师')
        subject = Subject.objects.create(name='信息技术', applicable_grades='1,2')
        assignment_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is ClassSubjectTeacher
        )

        workbook_bytes = self.build_workbook({
            assignment_sheet: [
                [school_class.name, subject.name, teacher.name, 'TRUE'],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(ClassSubjectTeacher.objects.count(), 0)
        self.assertEqual(response.data['results'][assignment_sheet]['created'], 0)
        self.assertEqual(response.data['results'][assignment_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('信息技术 不适用于 四1班 所在年级。', response.data['errors'][0])

    def test_import_rejects_invalid_teacher_combined_class_day(self):
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )

        workbook_bytes = self.build_workbook({
            teacher_sheet: [
                ['赵老师', None, None, 2, 'FALSE', None, None],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(Teacher.objects.filter(name='赵老师').count(), 0)
        self.assertEqual(response.data['results'][teacher_sheet]['created'], 0)
        self.assertEqual(response.data['results'][teacher_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('校本课程日期', response.data['errors'][0])

    def test_import_rejects_invalid_blocked_time_period_type(self):
        teacher = Teacher.objects.create(name='刘老师')
        blocked_time_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is TeacherBlockedTime
        )

        workbook_bytes = self.build_workbook({
            blocked_time_sheet: [
                [teacher.name, 1, 'night'],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(TeacherBlockedTime.objects.count(), 0)
        self.assertEqual(response.data['results'][blocked_time_sheet]['created'], 0)
        self.assertEqual(response.data['results'][blocked_time_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('时段', response.data['errors'][0])

    def test_import_rejects_when_existing_name_duplicates_make_lookup_ambiguous(self):
        Teacher.objects.create(name='王老师')
        Teacher.objects.create(name='王老师')
        blocked_time_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is TeacherBlockedTime
        )

        workbook_bytes = self.build_workbook({
            blocked_time_sheet: [
                ['王老师', 1, 'am'],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(TeacherBlockedTime.objects.count(), 0)
        self.assertEqual(response.data['results'][blocked_time_sheet]['created'], 0)
        self.assertEqual(response.data['results'][blocked_time_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('教师“王老师”存在多条记录', response.data['errors'][0])
        self.assertIn('王老师', response.data['errors'][0])

    def test_import_allows_duplicate_teacher_names_when_import_keys_are_present(self):
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )

        workbook_bytes = self.build_workbook({
            teacher_sheet: [
                ['周老师', None, None, None, 'FALSE', None, None, 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa'],
                ['周老师', None, None, None, 'FALSE', None, None, 'bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb'],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['committed'])
        self.assertEqual(Teacher.objects.filter(name='周老师').count(), 2)
        self.assertEqual(
            set(Teacher.objects.filter(name='周老师').values_list('import_key', flat=True)),
            {
                'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
                'bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb',
            }
        )
        self.assertEqual(response.data['results'][teacher_sheet]['created'], 2)
        self.assertEqual(response.data['results'][teacher_sheet]['updated'], 0)

    def test_import_resolves_duplicate_teacher_names_by_import_key(self):
        teacher_a = Teacher.objects.create(
            name='赵老师',
            import_key='11111111111111111111111111111111',
        )
        Teacher.objects.create(
            name='赵老师',
            import_key='22222222222222222222222222222222',
        )
        blocked_time_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is TeacherBlockedTime
        )

        workbook_bytes = self.build_workbook({
            blocked_time_sheet: [
                ['赵老师', 1, 'am', teacher_a.import_key],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['committed'])
        blocked_time = TeacherBlockedTime.objects.get()
        self.assertEqual(blocked_time.teacher_id, teacher_a.id)
        self.assertEqual(response.data['results'][blocked_time_sheet]['created'], 1)
        self.assertEqual(response.data['results'][blocked_time_sheet]['updated'], 0)

    def test_import_updates_existing_teacher_by_import_key(self):
        teacher = Teacher.objects.create(
            name='原教师',
            import_key='33333333333333333333333333333333',
        )
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )

        workbook_bytes = self.build_workbook({
            teacher_sheet: [
                ['新教师名', None, None, None, 'FALSE', None, None, teacher.import_key],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        teacher.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['committed'])
        self.assertEqual(teacher.name, '新教师名')
        self.assertEqual(response.data['results'][teacher_sheet]['created'], 0)
        self.assertEqual(response.data['results'][teacher_sheet]['updated'], 1)

    def test_import_updates_scheduler_settings(self):
        settings_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is SchedulerSettings
        )
        SchedulerSettings.get_settings()

        workbook_bytes = self.build_workbook({
            settings_sheet: [[
                '主题班会',
                '0,0;0,1',
                2,
                '0,1',
                1,
                'FALSE',
                2,
                11,
                12,
                13,
                4,
                15,
                16,
                17,
                18,
            ]]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        settings = SchedulerSettings.get_settings()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['committed'])
        self.assertEqual(settings.class_meeting_name, '主题班会')
        self.assertEqual(settings.combined_class_slots, '0,0;0,1')
        self.assertEqual(settings.solver_num_workers, 2)
        self.assertEqual(settings.h9_consecutive_forbidden, '0,1')
        self.assertFalse(settings.h14_homeroom_main_subject)
        self.assertEqual(settings.h15_teacher_max_main_subjects, 2)
        self.assertEqual(settings.s7_same_class_subject_switch_weight, 18)
        self.assertEqual(response.data['results'][settings_sheet]['created'], 0)
        self.assertEqual(response.data['results'][settings_sheet]['updated'], 1)

    def test_import_applies_scheduler_settings_before_teacher_qualification_filtering(self):
        settings_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is SchedulerSettings
        )
        subject_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Subject
        )
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )
        qualification_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is TeacherQualification
        )

        workbook_bytes = self.build_workbook({
            subject_sheet: [['主题班会']],
            settings_sheet: [['主题班会']],
            teacher_sheet: [['李老师']],
            qualification_sheet: [['李老师', '主题班会']],
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        settings = SchedulerSettings.get_settings()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['committed'])
        self.assertEqual(settings.class_meeting_name, '主题班会')
        self.assertEqual(TeacherQualification.objects.count(), 0)
        self.assertEqual(response.data['results'][qualification_sheet]['created'], 0)
        self.assertEqual(response.data['results'][qualification_sheet]['updated'], 0)

    def test_import_rejects_multiple_scheduler_settings_rows(self):
        settings_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is SchedulerSettings
        )
        original_settings = SchedulerSettings.get_settings()

        workbook_bytes = self.build_workbook({
            settings_sheet: [
                ['班会A', '1,4;1,5', 4, '1,2', 2, 'TRUE', 1, 10, 5, 2, 3, 8, 6, 5, 3],
                ['班会B', '3,4;3,5', 3, '0,1', 1, 'FALSE', 2, 9, 4, 1, 2, 7, 5, 4, 2],
            ]
        })
        upload = SimpleUploadedFile(
            'import.xlsx',
            workbook_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        request = self.factory.post('/api/data/import/', {'file': upload}, format='multipart')
        response = import_data(request)

        original_settings.refresh_from_db()
        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data['committed'])
        self.assertEqual(original_settings.class_meeting_name, '班会')
        self.assertEqual(response.data['results'][settings_sheet]['created'], 0)
        self.assertEqual(response.data['results'][settings_sheet]['updated'], 0)
        self.assertEqual(response.data['error_count'], 1)
        self.assertIn('只能包含一行参数设置', response.data['errors'][0])


class SchedulerSettingsSchemaMigrationTests(APITestCase):
    def test_normalize_scheduler_settings_schema_removes_legacy_required_columns(self):
        if connection.vendor != 'sqlite':
            self.skipTest('Only applicable to SQLite databases')

        migration_module = importlib.import_module(
            'core.migrations.0024_normalize_scheduler_settings_schema'
        )

        SchedulerSettings.objects.all().delete()

        column_definitions = ', '.join(
            f'"{column_name}" {column_sql}'
            for column_name, column_sql in migration_module.CURRENT_SCHEDULER_SETTINGS_COLUMNS
        )

        with connection.cursor() as cursor:
            cursor.execute('DROP TABLE IF EXISTS "core_schedulersettings_broken"')
            cursor.execute(
                'CREATE TABLE "core_schedulersettings_broken" ('
                f'{column_definitions}, '
                '"f2_enable_homeroom_main_subject" bool NOT NULL)'
            )
            cursor.execute('DROP TABLE "core_schedulersettings"')
            cursor.execute(
                'ALTER TABLE "core_schedulersettings_broken" '
                'RENAME TO "core_schedulersettings"'
            )
            before_columns = {
                row[1]
                for row in cursor.execute(
                    'PRAGMA table_info("core_schedulersettings")'
                ).fetchall()
            }

        self.assertIn('f2_enable_homeroom_main_subject', before_columns)

        schema_editor = SimpleNamespace(connection=connection)
        migration_module.normalize_scheduler_settings_schema(django_apps, schema_editor)

        with connection.cursor() as cursor:
            after_columns = {
                row[1]
                for row in cursor.execute(
                    'PRAGMA table_info("core_schedulersettings")'
                ).fetchall()
            }

        self.assertNotIn('f2_enable_homeroom_main_subject', after_columns)

        # 真实迁移链在 0024 之后会继续补充更新的字段（如 0026 的 h14、0027 的 h15）。
        # 规范化把表重建为 0024 时代的列集，这里手动补齐以便 ORM 正常访问单例表。
        with connection.cursor() as cursor:
            cursor.execute(
                'ALTER TABLE "core_schedulersettings" '
                'ADD COLUMN "h14_homeroom_main_subject" bool NOT NULL DEFAULT 1'
            )
            cursor.execute(
                'ALTER TABLE "core_schedulersettings" '
                'ADD COLUMN "h15_teacher_max_main_subjects" integer NOT NULL DEFAULT 1'
            )

        self.assertEqual(SchedulerSettings.get_settings().pk, 1)


class ExportDataTests(APITestCase):
    def setUp(self):
        self.factory = APIRequestFactory()

    def test_export_teacher_includes_combined_class_day(self):
        combined_group = CombinedClassGroup.objects.create(
            name='校本二组',
            import_key='cccccccccccccccccccccccccccccccc',
        )
        teacher = Teacher.objects.create(
            name='王老师',
            combined_class_group=combined_group,
            combined_class_day=1,
            import_key='dddddddddddddddddddddddddddddddd',
        )
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )

        request = self.factory.get('/api/data/export/')
        response = export_data(request)

        workbook = load_workbook(BytesIO(response.content))
        ws = workbook[teacher_sheet]
        headers = [cell.value for cell in ws[1]]
        data_row = [cell.value for cell in ws[2]]

        self.assertIn('校本课程日期(1=周二,3=周四，留空自动分配)', headers)
        self.assertIn(IMPORT_KEY_HEADER, headers)
        self.assertIn('校本课程分组导入键', headers)
        self.assertEqual(data_row[0], '王老师')
        self.assertEqual(data_row[2], combined_group.name)
        self.assertEqual(data_row[3], 1)
        self.assertEqual(data_row[7], teacher.import_key)
        self.assertEqual(data_row[9], combined_group.import_key)

    def test_export_allows_duplicate_teacher_names_and_includes_distinct_import_keys(self):
        Teacher.objects.create(
            name='李老师',
            import_key='eeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee',
        )
        Teacher.objects.create(
            name='李老师',
            import_key='ffffffffffffffffffffffffffffffff',
        )
        teacher_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is Teacher
        )

        request = self.factory.get('/api/data/export/')
        response = export_data(request)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook[teacher_sheet]
        exported_keys = {ws.cell(row=2, column=8).value, ws.cell(row=3, column=8).value}
        self.assertEqual(
            exported_keys,
            {'eeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee', 'ffffffffffffffffffffffffffffffff'}
        )

    def test_export_includes_scheduler_settings_sheet(self):
        settings_sheet = next(
            name for name, config in SHEET_CONFIG.items()
            if config['model'] is SchedulerSettings
        )
        settings = SchedulerSettings.get_settings()
        settings.class_meeting_name = '主题班会'
        settings.combined_class_slots = '0,0;0,1'
        settings.solver_num_workers = 2
        settings.save()

        request = self.factory.get('/api/data/export/')
        response = export_data(request)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn(settings_sheet, workbook.sheetnames)
        ws = workbook[settings_sheet]
        headers = [cell.value for cell in ws[1]]
        data_row = [cell.value for cell in ws[2]]

        self.assertEqual(headers[0], '班会课程名称')
        self.assertEqual(headers[2], '求解器线程数')
        self.assertEqual(data_row[0], '主题班会')
        self.assertEqual(data_row[1], '0,0;0,1')
        self.assertEqual(data_row[2], 2)
