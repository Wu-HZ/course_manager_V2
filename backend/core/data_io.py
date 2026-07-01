"""
数据导入导出功能
Excel 格式，每个数据类型一个 Sheet
"""
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .models import (
    ClassSubjectTeacher,
    CombinedClassGroup,
    Location,
    ScheduleLock,
    SchedulerSettings,
    School,
    SchoolClass,
    Subject,
    Teacher,
    TeacherBlockedTime,
    TeacherQualification,
    TravelGroup,
    get_assignment_subject_validation_error,
    get_qualification_subject_queryset,
    is_subject_qualification_managed,
)
from .school_utils import get_request_school

IMPORT_KEY_HEADER = '导入键(系统生成，请勿修改)'
BASE_MODELS = {
    TravelGroup,
    Subject,
    CombinedClassGroup,
    Teacher,
    SchoolClass,
    Location,
}


SHEET_CONFIG = {
    '送教分组': {
        'model': TravelGroup,
        'fields': ['name', 'day_off', 'import_key'],
        'headers': ['分组名称', '禁排日(0周一~4周五)', IMPORT_KEY_HEADER],
    },
    '校本课程分组': {
        'model': CombinedClassGroup,
        'fields': ['name', 'import_key'],
        'headers': ['分组名称', IMPORT_KEY_HEADER],
    },
    '场地': {
        'model': Location,
        'fields': ['name', 'location_type', 'capacity', 'import_key'],
        'headers': ['场地名称', '类型(CLASSROOM/PLAYGROUND/LAB/HOME_EC)', '容量', IMPORT_KEY_HEADER],
    },
    '课程': {
        'model': Subject,
        'fields': [
            'name', 'weekly_hours', 'is_main_subject', 'max_teacher_classes',
            'is_am_preferred', 'allow_consecutive', 'max_daily_limit',
            'location_type', 'is_combined_class', 'applicable_grades',
            'avoid_first_period', 'import_key',
        ],
        'headers': [
            '课程名称', '周课时', '主课(TRUE/FALSE)', '单师班数(1-5)',
            '优先上午(TRUE/FALSE)', '允许连堂(TRUE/FALSE)', '单日上限',
            '场地类型', '是否合班课(TRUE/FALSE)', '适用年级(如 1,2,3)',
            '避免第一节(TRUE/FALSE)', IMPORT_KEY_HEADER,
        ],
    },
    '教师': {
        'model': Teacher,
        'fields': [
            'name', 'travel_group__name', 'combined_class_group__name',
            'combined_class_day', 'exclude_from_combined',
            'min_weekly_hours', 'max_weekly_hours', 'import_key',
            'travel_group__import_key', 'combined_class_group__import_key',
        ],
        'headers': [
            '姓名', '送教分组名称', '校本课程分组名称',
            '校本课程日期(1=周二,3=周四，留空自动分配)',
            '不参与校本课程(TRUE/FALSE)', '周课时下限(留空不限)',
            '周课时上限(留空不限)', IMPORT_KEY_HEADER,
            '送教分组导入键', '校本课程分组导入键',
        ],
        'fk_fields': {
            'travel_group__name': ('travel_group', TravelGroup, 'name'),
            'combined_class_group__name': ('combined_class_group', CombinedClassGroup, 'name'),
            'travel_group__import_key': ('travel_group', TravelGroup, 'import_key'),
            'combined_class_group__import_key': ('combined_class_group', CombinedClassGroup, 'import_key'),
        },
    },
    '班级': {
        'model': SchoolClass,
        'fields': ['name', 'grade', 'homeroom_teacher__name', 'import_key', 'homeroom_teacher__import_key'],
        'headers': ['班级名称', '年级', '班主任姓名', IMPORT_KEY_HEADER, '班主任导入键'],
        'fk_fields': {
            'homeroom_teacher__name': ('homeroom_teacher', Teacher, 'name'),
            'homeroom_teacher__import_key': ('homeroom_teacher', Teacher, 'import_key'),
        },
    },
    '教师资质': {
        'model': TeacherQualification,
        'fields': ['teacher__name', 'subject__name', 'teacher__import_key', 'subject__import_key'],
        'headers': ['教师姓名', '课程名称', '教师导入键', '课程导入键'],
        'fk_fields': {
            'teacher__name': ('teacher', Teacher, 'name'),
            'subject__name': ('subject', Subject, 'name'),
            'teacher__import_key': ('teacher', Teacher, 'import_key'),
            'subject__import_key': ('subject', Subject, 'import_key'),
        },
    },
    '授课分配': {
        'model': ClassSubjectTeacher,
        'fields': [
            'school_class__name', 'subject__name', 'teacher__name', 'is_manual',
            'school_class__import_key', 'subject__import_key', 'teacher__import_key',
        ],
        'headers': [
            '班级名称', '课程名称', '教师姓名', '手动指定(TRUE/FALSE)',
            '班级导入键', '课程导入键', '教师导入键',
        ],
        'fk_fields': {
            'school_class__name': ('school_class', SchoolClass, 'name'),
            'subject__name': ('subject', Subject, 'name'),
            'teacher__name': ('teacher', Teacher, 'name'),
            'school_class__import_key': ('school_class', SchoolClass, 'import_key'),
            'subject__import_key': ('subject', Subject, 'import_key'),
            'teacher__import_key': ('teacher', Teacher, 'import_key'),
        },
    },
    '课表锁定': {
        'model': ScheduleLock,
        'fields': [
            'school_class__name', 'subject__name', 'teacher__name', 'day', 'period',
            'school_class__import_key', 'subject__import_key', 'teacher__import_key',
        ],
        'headers': [
            '班级名称', '课程名称', '教师姓名(可空)', '星期(0-4)', '节次(0-5)',
            '班级导入键', '课程导入键', '教师导入键',
        ],
        'fk_fields': {
            'school_class__name': ('school_class', SchoolClass, 'name'),
            'subject__name': ('subject', Subject, 'name'),
            'teacher__name': ('teacher', Teacher, 'name'),
            'school_class__import_key': ('school_class', SchoolClass, 'import_key'),
            'subject__import_key': ('subject', Subject, 'import_key'),
            'teacher__import_key': ('teacher', Teacher, 'import_key'),
        },
    },
    '教师禁排日': {
        'model': TeacherBlockedTime,
        'fields': ['teacher__name', 'day', 'period_type', 'teacher__import_key'],
        'headers': ['教师姓名', '星期(0-4)', '时段(am/pm/all)', '教师导入键'],
        'fk_fields': {
            'teacher__name': ('teacher', Teacher, 'name'),
            'teacher__import_key': ('teacher', Teacher, 'import_key'),
        },
    },
    '排课参数': {
        'model': SchedulerSettings,
        'fields': [
            'class_meeting_name',
            'combined_class_slots',
            'solver_num_workers',
            'h9_consecutive_forbidden',
            'h11_teacher_class_daily_max',
            'h14_homeroom_main_subject',
            'h15_teacher_max_main_subjects',
            's1_am_preference_weight',
            's2_consecutive_weight',
            's3_distribution_weight',
            's4_teacher_daily_threshold',
            's4_teacher_daily_weight',
            's5_avoid_first_weight',
            's6_subject_switch_weight',
            's7_same_class_subject_switch_weight',
        ],
        'headers': [
            '班会课程名称',
            '合班课时段',
            '求解器线程数',
            '教师禁止跨边界连续上课',
            '教师同班单日上限',
            '班主任必须担任主课(TRUE/FALSE)',
            '单师最多主课数',
            '上午优先权重',
            '连堂偏好权重',
            '分布均匀权重',
            '教师日负荷阈值',
            '教师日负荷权重',
            '避免第一节权重',
            '换班惩罚权重',
            '同班换科惩罚权重',
        ],
    },
}

EXPORT_ORDER = ['送教分组', '校本课程分组', '场地', '课程', '教师', '教师禁排日', '班级', '教师资质', '授课分配', '课表锁定', '排课参数']
IMPORT_ORDER = ['送教分组', '校本课程分组', '场地', '课程', '排课参数', '教师', '教师禁排日', '班级', '教师资质', '授课分配', '课表锁定']


def style_header(ws):
    """设置表头样式"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


def build_import_response(results, errors, committed, error_message=None):
    """统一构造导入响应，便于前端区分是否已落库。"""
    payload = {
        'results': results,
        'errors': errors[:20],
        'error_count': len(errors),
        'committed': committed,
    }

    if committed:
        return Response(payload, status=200)

    payload['error'] = error_message or '导入失败，已回滚，本次未写入任何数据。'
    payload['results'] = {
        sheet_name: {'created': 0, 'updated': 0}
        for sheet_name in results
    }
    return Response(payload, status=400)


def get_export_value(obj, field):
    """获取导出字段值。"""
    if '__' in field:
        value = obj
        for part in field.split('__'):
            value = getattr(value, part, None) if value else None
    else:
        value = getattr(obj, field, None)

    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    return value


def normalize_cell_value(value):
    """统一处理 Excel 单元格值。"""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.upper() == 'TRUE':
            return True
        if stripped.upper() == 'FALSE':
            return False
        return stripped
    return value


def is_base_model(model):
    return model in BASE_MODELS


def get_model_lookup_kwargs(model, data):
    """返回导入时定位现有记录所需的查询条件。"""
    if is_base_model(model):
        if data.get('import_key'):
            return {'import_key': data['import_key']}
        if data.get('name'):
            return {'name': data['name']}
        return None

    if model == TeacherQualification:
        return {'teacher': data.get('teacher'), 'subject': data.get('subject')}
    if model == ClassSubjectTeacher:
        return {'school_class': data.get('school_class'), 'subject': data.get('subject')}
    if model == ScheduleLock:
        return {
            'school_class': data.get('school_class'),
            'day': data.get('day'),
            'period': data.get('period'),
        }
    if model == TeacherBlockedTime:
        return {
            'teacher': data.get('teacher'),
            'day': data.get('day'),
            'period_type': data.get('period_type'),
        }
    if model == SchedulerSettings:
        return {'school': data.get('school')}
    return None


def describe_lookup_conflict(model, lookup_kwargs):
    if not lookup_kwargs:
        return f'无法确定要更新的{model._meta.verbose_name}。'

    if 'name' in lookup_kwargs:
        return (
            f'{model._meta.verbose_name}“{lookup_kwargs["name"]}”存在多条记录，'
            '请使用带导入键的新导出模板。'
        )

    if 'import_key' in lookup_kwargs:
        return (
            f'{model._meta.verbose_name}导入键“{lookup_kwargs["import_key"]}”对应了多条记录。'
        )

    return f'存在多条匹配的{model._meta.verbose_name}记录。'


def get_existing_import_instance(model, lookup_kwargs):
    """获取将被本次导入更新的已有记录。"""
    if not lookup_kwargs:
        return None, None

    try:
        return model.objects.get(**lookup_kwargs), None
    except model.DoesNotExist:
        return None, None
    except model.MultipleObjectsReturned:
        return None, describe_lookup_conflict(model, lookup_kwargs)


def format_model_validation_error(model, error):
    """把 Django ValidationError 格式化成更适合导入结果展示的文本。"""
    if hasattr(error, 'message_dict'):
        messages = []
        for field_name, field_errors in error.message_dict.items():
            if field_name == '__all__':
                messages.extend(str(message) for message in field_errors)
                continue
            try:
                label = model._meta.get_field(field_name).verbose_name
            except Exception:
                label = field_name
            messages.extend(f'{label}: {message}' for message in field_errors)
        return '；'.join(messages)

    return '；'.join(str(message) for message in error.messages)


def validate_import_instance(model, data, existing_instance=None):
    """在真正写库前执行模型层校验。"""
    instance = existing_instance or model()
    for key, value in data.items():
        setattr(instance, key, value)
    instance.full_clean()


def collect_row_data(model, fields, fk_fields, row):
    """解析单行数据，拆分普通字段和待解析外键。"""
    data = {}
    pending_fks = {}

    for index, field in enumerate(fields):
        value = normalize_cell_value(row[index] if index < len(row) else None)

        if field in fk_fields:
            fk_field_name, fk_model, fk_lookup = fk_fields[field]
            pending_entry = pending_fks.setdefault(
                fk_field_name,
                {'model': fk_model, 'values': {}},
            )
            pending_entry['values'][fk_lookup] = value
            continue

        if field == 'import_key':
            if value:
                data[field] = value
            continue

        if value is None:
            model_field = model._meta.get_field(field)
            if model_field.has_default():
                value = model_field.default

        data[field] = value

    return data, pending_fks


def resolve_foreign_key(target_field, fk_model, lookup_values):
    """按导入键优先、名称兜底的顺序解析外键。"""
    import_key = lookup_values.get('import_key')
    name = lookup_values.get('name')

    if import_key:
        try:
            return fk_model.objects.get(import_key=import_key), None
        except fk_model.DoesNotExist:
            return None, f'找不到{fk_model._meta.verbose_name}，导入键为“{import_key}”。'
        except fk_model.MultipleObjectsReturned:
            return None, f'{fk_model._meta.verbose_name}导入键“{import_key}”对应了多条记录。'

    if name:
        try:
            return fk_model.objects.get(name=name), None
        except fk_model.DoesNotExist:
            return None, f'找不到{fk_model._meta.verbose_name}“{name}”。'
        except fk_model.MultipleObjectsReturned:
            return None, (
                f'{fk_model._meta.verbose_name}“{name}”存在多条记录，'
                '请使用带导入键的新导出模板。'
            )

    return None, None


def resolve_pending_foreign_keys(data, pending_fks):
    """把待解析外键填充到 data 中。"""
    for target_field, entry in pending_fks.items():
        fk_obj, error = resolve_foreign_key(
            target_field,
            entry['model'],
            entry['values'],
        )
        if error:
            return error
        data[target_field] = fk_obj
    return None


def save_import_instance(model, data, existing_instance=None):
    """按是否已有记录执行创建或更新。"""
    if existing_instance is None:
        obj = model.objects.create(**data)
        return obj, True

    for key, value in data.items():
        setattr(existing_instance, key, value)
    existing_instance.save()
    return existing_instance, False


@api_view(['GET'])
def export_data(request):
    """导出当前学校的所有数据到 Excel。"""
    school = get_request_school(request)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in EXPORT_ORDER:
        config = SHEET_CONFIG[sheet_name]
        model = config['model']
        fields = config['fields']
        headers = config['headers']

        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        style_header(ws)

        if model == SchedulerSettings:
            queryset = [SchedulerSettings.get_settings(school)]
        elif hasattr(model, 'school') and hasattr(model._meta, 'get_field'):
            try:
                model._meta.get_field('school')
                queryset = model.objects.filter(school=school)
            except Exception:
                queryset = model.objects.all()
        else:
            queryset = model.objects.all()

        if model == TeacherQualification:
            queryset = queryset.filter(subject__in=get_qualification_subject_queryset(school))

        for obj in queryset:
            ws.append([get_export_value(obj, field) for field in fields])

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="course_manager_data.xlsx"'
    return response


@api_view(['POST'])
def import_data(request):
    """从 Excel 导入数据到当前学校。"""
    school = get_request_school(request)
    file = request.FILES.get('file')
    if not file:
        return Response({'error': '请上传文件'}, status=400)

    try:
        wb = load_workbook(file)
    except Exception as exc:
        return Response({'error': f'无法读取 Excel 文件: {exc}'}, status=400)

    results = {
        sheet_name: {'created': 0, 'updated': 0}
        for sheet_name in EXPORT_ORDER
        if sheet_name in wb.sheetnames
    }
    errors = []

    with transaction.atomic():
        for sheet_name in IMPORT_ORDER:
            if sheet_name not in wb.sheetnames:
                continue

            config = SHEET_CONFIG[sheet_name]
            model = config['model']
            fields = config['fields']
            fk_fields = config.get('fk_fields', {})
            ws = wb[sheet_name]

            created = 0
            updated = 0
            singleton_row_imported = False

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    if model == SchedulerSettings and singleton_row_imported:
                        errors.append(f'{sheet_name} 第{row_idx}行: 该工作表只能包含一行参数设置。')
                        continue

                    data, pending_fks = collect_row_data(model, fields, fk_fields, row)
                    if not data and not pending_fks:
                        continue

                    fk_error = resolve_pending_foreign_keys(data, pending_fks)
                    if fk_error:
                        errors.append(f'{sheet_name} 第{row_idx}行: {fk_error}')
                        continue

                    # 为需要 school 的模型自动注入 school
                    if model != SchedulerSettings and hasattr(model, 'school'):
                        try:
                            model._meta.get_field('school')
                            data['school'] = school
                        except Exception:
                            pass

                    lookup_kwargs = get_model_lookup_kwargs(model, data)
                    # 查找时也限定学校
                    if lookup_kwargs and 'school' not in lookup_kwargs and hasattr(model, 'school'):
                        try:
                            model._meta.get_field('school')
                            lookup_kwargs['school'] = school
                        except Exception:
                            pass
                    existing_instance, lookup_error = get_existing_import_instance(model, lookup_kwargs)
                    if lookup_error:
                        errors.append(f'{sheet_name} 第{row_idx}行: {lookup_error}')
                        continue

                    if model == TeacherQualification:
                        if not is_subject_qualification_managed(data.get('subject')):
                            continue

                    if model == ClassSubjectTeacher:
                        validation_error = get_assignment_subject_validation_error(
                            data.get('school_class'),
                            data.get('subject'),
                        )
                        if validation_error:
                            errors.append(f'{sheet_name} 第{row_idx}行: {validation_error}')
                            continue

                    if model == SchedulerSettings:
                        # SchedulerSettings: 确保使用当前学校的设置
                        data['school'] = school

                    try:
                        validate_import_instance(model, data, existing_instance)
                    except ValidationError as exc:
                        errors.append(
                            f'{sheet_name} 第{row_idx}行: '
                            f'{format_model_validation_error(model, exc)}'
                        )
                        continue

                    _, is_created = save_import_instance(model, data, existing_instance)
                    if model == SchedulerSettings:
                        singleton_row_imported = True
                    if is_created:
                        created += 1
                    else:
                        updated += 1

                except Exception as exc:
                    errors.append(f'{sheet_name} 第{row_idx}行: {exc}')

            results[sheet_name] = {'created': created, 'updated': updated}

        if errors:
            transaction.set_rollback(True)
            return build_import_response(results, errors, committed=False)

    return build_import_response(results, errors, committed=True)
